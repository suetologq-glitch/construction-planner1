from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from datetime import datetime, timedelta, date
from models import db, Project, Stage, Dependency, Resource, Assignment, Photo, PurchaseOrder, PurchaseItem, Act, ActItem, Contract, ContractWorkItem, User, ProjectTemplate, Notification, ProjectTeam
from plan_logic import calculate_dates, find_critical_path, predict_completion_date
from io import BytesIO
import openpyxl
from openpyxl.styles import Font
import os
import requests
import csv
from io import StringIO
from docx import Document
from docx.shared import Inches, Pt
from docx.enum.text import WD_ALIGN_PARAGRAPH
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from functools import wraps
import json
from werkzeug.utils import secure_filename

app = Flask(__name__)

# === НАСТРОЙКА БАЗЫ ДАННЫХ ===
basedir = os.path.abspath(os.path.dirname(__file__))
db_path = os.path.join(basedir, 'construction.db')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + db_path
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'your-secret-key-here'

print(f"📁 База данных сохранена в: {db_path}")

db.init_app(app)

# API ключ для OpenWeatherMap (ваш рабочий ключ)
WEATHER_API_KEY = "45e15fb41b97d5a968cae6fdc72e88c5"

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

UPLOAD_FOLDER = os.path.join('static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

with app.app_context():
    db.create_all()
    print("✅ Таблицы базы данных проверены/созданы.")

# ========== ДЕКОРАТОРЫ ДОСТУПА ==========
def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if current_user.role not in roles:
                flash('У вас недостаточно прав для доступа к этой странице.', 'error')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def check_project_access(project_id):
    project = Project.query.get(project_id)
    if not project:
        return False
    if project.user_id == current_user.id:
        return True
    if current_user.is_admin():
        return True
    team_member = ProjectTeam.query.filter_by(project_id=project_id, user_id=current_user.id).first()
    return team_member is not None

def project_access_required(f):
    @wraps(f)
    def decorated_function(project_id, *args, **kwargs):
        if not check_project_access(project_id):
            flash('У вас нет доступа к этому проекту.', 'error')
            return redirect(url_for('index'))
        return f(project_id, *args, **kwargs)
    return decorated_function

@app.template_filter('strftime')
def _jinja2_filter_datetime(date, fmt=None):
    if date is None:
        return ''
    return date.strftime('%Y-%m-%d')

# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========
def get_overdue_stages():
    now = datetime.now().date()
    overdue_stages = []
    all_stages = Stage.query.all()
    for stage in all_stages:
        if stage.percent_complete < 100:
            if stage.custom_start_date and stage.custom_end_date:
                planned_end = stage.custom_end_date
            else:
                dates = calculate_dates(stage.project, db)
                planned_end = dates.get(stage.id, (None, None))[1]
            if planned_end and planned_end < now:
                overdue_stages.append({
                    'stage': stage,
                    'project': stage.project,
                    'planned_end': planned_end,
                    'days_overdue': (now - planned_end).days
                })
    return overdue_stages

def get_weather_by_coords(lat, lon):
    try:
        url = f"http://api.openweathermap.org/data/2.5/weather?lat={lat}&lon={lon}&units=metric&lang=ru&appid={WEATHER_API_KEY}"
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            data = response.json()
            return {
                'temp': round(data['main']['temp']),
                'feels_like': round(data['main']['feels_like']),
                'humidity': data['main']['humidity'],
                'description': data['weather'][0]['description'],
                'icon': data['weather'][0]['icon'],
                'wind_speed': data['wind']['speed'],
                'city': data.get('name', 'Объект'),
                'lat': lat,
                'lon': lon
            }
        return None
    except Exception as e:
        print(f"Ошибка получения погоды: {e}")
        return None

def calculate_s_curve(project):
    stages = Stage.query.filter_by(project_id=project.id).all()
    if not stages:
        return {'dates': [], 'planned': [], 'actual': []}
    dates = {}
    for s in stages:
        if s.custom_start_date and s.custom_end_date:
            dates[s.id] = (s.custom_start_date, s.custom_end_date)
        else:
            dates.update(calculate_dates(project, db))
    if not dates:
        return {'dates': [], 'planned': [], 'actual': []}
    min_date = min(start for start, _ in dates.values())
    max_date = max(end for _, end in dates.values())
    progress_by_day = {}
    current = min_date
    while current <= max_date:
        progress_by_day[current] = {'planned': 0, 'actual': 0}
        current += timedelta(days=1)
    for stage in stages:
        if stage.id not in dates:
            continue
        start_date, end_date = dates[stage.id]
        duration = (end_date - start_date).days + 1
        planned_daily = stage.percent_complete / duration if duration > 0 else 0
        actual_daily = stage.percent_complete / duration if duration > 0 else 0
        current = start_date
        while current <= end_date:
            if current in progress_by_day:
                progress_by_day[current]['planned'] += planned_daily
                if stage.actual_end_date and current <= stage.actual_end_date:
                    progress_by_day[current]['actual'] += actual_daily
            current += timedelta(days=1)
    dates_list = sorted(progress_by_day.keys())
    planned_curve = []
    actual_curve = []
    planned_sum = 0
    actual_sum = 0
    for d in dates_list:
        planned_sum += progress_by_day[d]['planned']
        actual_sum += progress_by_day[d]['actual']
        planned_curve.append(min(planned_sum, 100))
        actual_curve.append(min(actual_sum, 100))
    return {
        'dates': [d.strftime('%d.%m.%Y') for d in dates_list],
        'planned': planned_curve,
        'actual': actual_curve
    }

def get_materials_for_stage(stage_name, budget):
    stage_lower = stage_name.lower()
    materials = []
    if 'фундамент' in stage_lower:
        materials.append({'name': 'Бетон М300', 'quantity': budget / 5000, 'unit': 'м³', 'unit_price': 5000, 'total_price': budget * 0.6, 'supplier': 'БетонСтрой'})
        materials.append({'name': 'Арматура', 'quantity': budget / 20000, 'unit': 'т', 'unit_price': 20000, 'total_price': budget * 0.4, 'supplier': 'МеталлТорг'})
    elif 'стены' in stage_lower or 'кирпич' in stage_lower:
        materials.append({'name': 'Газоблок', 'quantity': budget / 6000, 'unit': 'м³', 'unit_price': 6000, 'total_price': budget * 0.7, 'supplier': 'СтройБлок'})
        materials.append({'name': 'Клей для блоков', 'quantity': budget / 20000, 'unit': 'мешок', 'unit_price': 400, 'total_price': budget * 0.3, 'supplier': 'СтройБлок'})
    elif 'кровля' in stage_lower:
        materials.append({'name': 'Металлочерепица', 'quantity': budget / 800, 'unit': 'м²', 'unit_price': 800, 'total_price': budget * 0.8, 'supplier': 'КровляПро'})
        materials.append({'name': 'Гидроизоляция', 'quantity': budget / 500, 'unit': 'рулон', 'unit_price': 2000, 'total_price': budget * 0.2, 'supplier': 'КровляПро'})
    elif 'фасад' in stage_lower:
        materials.append({'name': 'Утеплитель', 'quantity': budget / 400, 'unit': 'м²', 'unit_price': 400, 'total_price': budget * 0.6, 'supplier': 'ТеплоСтрой'})
        materials.append({'name': 'Штукатурка фасадная', 'quantity': budget / 800, 'unit': 'мешок', 'unit_price': 500, 'total_price': budget * 0.4, 'supplier': 'ТеплоСтрой'})
    elif 'штукатурка' in stage_lower:
        materials.append({'name': 'Штукатурная смесь', 'quantity': budget / 400, 'unit': 'мешок', 'unit_price': 400, 'total_price': budget, 'supplier': 'ОтделкаСтрой'})
    elif 'стяжка' in stage_lower:
        materials.append({'name': 'Цемент', 'quantity': budget / 500, 'unit': 'мешок', 'unit_price': 500, 'total_price': budget * 0.6, 'supplier': 'ЦементТрейд'})
        materials.append({'name': 'Песок', 'quantity': budget / 2000, 'unit': 'м³', 'unit_price': 2000, 'total_price': budget * 0.4, 'supplier': 'ЦементТрейд'})
    elif 'отделка' in stage_lower or 'покраска' in stage_lower:
        materials.append({'name': 'Краска', 'quantity': budget / 800, 'unit': 'ведро', 'unit_price': 800, 'total_price': budget * 0.5, 'supplier': 'КраскиМир'})
        materials.append({'name': 'Обои', 'quantity': budget / 500, 'unit': 'рулон', 'unit_price': 500, 'total_price': budget * 0.5, 'supplier': 'КраскиМир'})
    elif 'плитка' in stage_lower or 'пол' in stage_lower:
        materials.append({'name': 'Плитка', 'quantity': budget / 1000, 'unit': 'м²', 'unit_price': 1000, 'total_price': budget, 'supplier': 'ПлиткаСтрой'})
    elif 'окна' in stage_lower or 'двери' in stage_lower:
        materials.append({'name': 'Окна ПВХ', 'quantity': 1, 'unit': 'компл', 'unit_price': budget, 'total_price': budget, 'supplier': 'ОкнаПро'})
    else:
        materials.append({'name': 'Строительные материалы', 'quantity': 1, 'unit': 'компл', 'unit_price': budget, 'total_price': budget, 'supplier': 'СтройМаркет'})
    return materials

def generate_stages_by_budget(budget, quality, house_area=None, land_area=None):
    quality_mult = {'econom': 0.8, 'standard': 1.0, 'premium': 1.3}
    mult = quality_mult.get(quality, 1.0)
    if house_area is None: house_area = 100
    if land_area is None: land_area = 600
    area_factor = house_area / 100.0
    land_factor = land_area / 600.0
    stages_config = [
        ('Геодезические работы', 2, None, 'Геодезист', 1, False, True),
        ('Подготовка участка', 3, 'Геодезические работы', 'Бригада 1', 2, False, True),
        ('Фундамент', 10, 'Подготовка участка', 'Бригада 1', 15, True, False),
        ('Стены', 12, 'Фундамент', 'Бригада 2', 20, True, False),
        ('Кровля', 8, 'Стены', 'Бригада 3', 12, True, False),
        ('Фасад', 10, 'Стены', 'Бригада 4', 10, True, False),
        ('Окна и двери', 5, 'Стены', 'Бригада 5', 5, True, False),
        ('Электрика', 8, 'Стены', 'Бригада 6', 6, True, False),
        ('Сантехника', 6, 'Стены', 'Бригада 7', 5, True, False),
        ('Отопление', 6, 'Стены', 'Бригада 7', 5, True, False),
        ('Штукатурка', 8, 'Электрика, Сантехника, Отопление', 'Бригада 8', 6, True, False),
        ('Стяжка', 5, 'Штукатурка', 'Бригада 8', 3, True, False),
        ('Чистовая отделка', 15, 'Стяжка', 'Бригада 9', 12, True, False),
        ('Благоустройство', 7, 'Фасад', 'Бригада 10', 5, False, True),
        ('Уборка и сдача', 3, 'Чистовая отделка, Благоустройство', 'Бригада 11', 2, True, True)
    ]
    stages = []
    total_weight = 0
    temp_stages = []
    for name, base_dur, depends_on, resources, percent, depends_house, depends_land in stages_config:
        cost_scale = 1.0
        if depends_house: cost_scale *= area_factor
        if depends_land: cost_scale *= land_factor
        duration = max(1, int(base_dur * (area_factor if depends_house else 1) * (land_factor if depends_land else 1)))
        stage_budget_share = percent / 100.0 * cost_scale
        temp_stages.append({
            'name': name, 'duration': duration, 'depends_on': depends_on,
            'resources': resources, 'share': stage_budget_share
        })
        total_weight += stage_budget_share
    for s in temp_stages:
        stage_budget = budget * (s['share'] / total_weight) * mult
        stages.append({
            'name': s['name'], 'duration': s['duration'],
            'depends_on': s['depends_on'], 'resources': s['resources'],
            'budget': round(stage_budget, 0)
        })
    return stages

def create_notification(user_id, title, message, link=None):
    notif = Notification(user_id=user_id, title=title, message=message, link=link)
    db.session.add(notif)
    db.session.commit()
    print(f"🔔 Уведомление для user {user_id}: {title}")

# ========== АУТЕНТИФИКАЦИЯ ==========
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        confirm = request.form['confirm_password']
        if password != confirm:
            flash('Пароли не совпадают', 'error')
            return redirect(url_for('register'))
        if len(password) < 6:
            flash('Пароль должен быть не менее 6 символов', 'error')
            return redirect(url_for('register'))
        if User.query.filter_by(username=username).first():
            flash('Имя пользователя уже занято', 'error')
            return redirect(url_for('register'))
        if User.query.filter_by(email=email).first():
            flash('Email уже зарегистрирован', 'error')
            return redirect(url_for('register'))
        role = 'admin' if User.query.count() == 0 else 'user'
        user = User(username=username, email=email, role=role)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash('Регистрация успешна! Войдите в систему.', 'success')
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(password):
            login_user(user)
            flash(f'Добро пожаловать, {user.username}!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Неверный email или пароль', 'error')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Вы вышли из системы', 'success')
    return redirect(url_for('login'))

# ========== ГЛАВНАЯ / ПРОЕКТЫ ==========
@app.route('/')
@login_required
def index():
    owned = Project.query.filter_by(user_id=current_user.id).all()
    team_projects = db.session.query(Project).join(ProjectTeam).filter(ProjectTeam.user_id == current_user.id).all()
    projects = list(set(owned + team_projects))
    templates = ProjectTemplate.query.filter_by(user_id=current_user.id).all()
    return render_template('index.html', projects=projects, templates=templates)

@app.route('/add_project', methods=['POST'])
@login_required
def add_project():
    name = request.form['name']
    start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date()
    address = request.form.get('address', '')
    try:
        lat = float(request.form.get('latitude', 55.751244))
        lon = float(request.form.get('longitude', 37.618423))
    except:
        lat, lon = 55.751244, 37.618423
    project = Project(name=name, start_date=start_date, address=address,
                      latitude=lat, longitude=lon, user_id=current_user.id)
    db.session.add(project)
    db.session.commit()
    flash('Проект создан', 'success')
    return redirect(url_for('index'))

@app.route('/delete_project/<int:project_id>')
@login_required
def delete_project(project_id):
    if not check_project_access(project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    db.session.delete(project)
    db.session.commit()
    flash('Проект удалён', 'success')
    return redirect(url_for('index'))

@app.route('/update_project_location/<int:project_id>', methods=['POST'])
@login_required
def update_project_location(project_id):
    if not check_project_access(project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    project = Project.query.get_or_404(project_id)
    data = request.get_json()
    project.latitude = data['latitude']
    project.longitude = data['longitude']
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/weather')
def api_weather():
    lat = request.args.get('lat', type=float)
    lon = request.args.get('lon', type=float)
    if lat is None or lon is None:
        return jsonify({'success': False, 'error': 'No coordinates'})
    weather = get_weather_by_coords(lat, lon)
    if weather:
        return jsonify({'success': True, **weather})
    return jsonify({'success': False, 'error': 'Weather not found'})

# ========== УВЕДОМЛЕНИЯ ==========
@app.route('/notifications')
@login_required
def notifications():
    notifs = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.created_at.desc()).all()
    return render_template('notifications.html', notifications=notifs)

@app.route('/notifications/read/<int:notif_id>')
@login_required
def mark_notification_read(notif_id):
    notif = Notification.query.get_or_404(notif_id)
    if notif.user_id != current_user.id:
        return jsonify({'success': False})
    notif.is_read = True
    db.session.commit()
    return jsonify({'success': True})

# ========== АДМИНИСТРИРОВАНИЕ ПОЛЬЗОВАТЕЛЕЙ ==========
@app.route('/admin/users')
@login_required
@role_required('admin')
def admin_users():
    users = User.query.all()
    return render_template('admin_users.html', users=users)

@app.route('/admin/user/<int:user_id>/role', methods=['POST'])
@login_required
@role_required('admin')
def update_user_role(user_id):
    user = User.query.get_or_404(user_id)
    new_role = request.form.get('role')
    if new_role in ['admin', 'manager', 'foreman', 'supplier', 'user']:
        user.role = new_role
        db.session.commit()
        flash(f'Роль пользователя {user.username} изменена на {new_role}', 'success')
    return redirect(url_for('admin_users'))

# ========== КОМАНДА ПРОЕКТА ==========
@app.route('/project/<int:project_id>/team', methods=['GET', 'POST'])
@login_required
def project_team(project_id):
    if not check_project_access(project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    if request.method == 'POST':
        email = request.form.get('email')
        role = request.form.get('role')
        user = User.query.filter_by(email=email).first()
        if user:
            existing = ProjectTeam.query.filter_by(project_id=project_id, user_id=user.id).first()
            if not existing:
                team_member = ProjectTeam(project_id=project_id, user_id=user.id, role=role)
                db.session.add(team_member)
                db.session.commit()
                flash(f'{user.username} добавлен в команду как {role}', 'success')
        else:
            flash('Пользователь с таким email не найден', 'error')
    team = ProjectTeam.query.filter_by(project_id=project_id).all()
    return render_template('project_team.html', project=project, team=team)

# ========== ПРОСМОТР ПРОЕКТА ==========
@app.route('/project/<int:project_id>')
@login_required
def project_view(project_id):
    if not check_project_access(project_id):
        flash('У вас нет доступа к этому проекту', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    stages = Stage.query.filter_by(project_id=project_id).all()
    dates = {}
    for s in stages:
        if s.custom_start_date and s.custom_end_date:
            dates[s.id] = (s.custom_start_date, s.custom_end_date)
        else:
            dates.update(calculate_dates(project, db))
    critical_stages = find_critical_path(project, db)
    assignments = {}
    for stage in stages:
        ass = Assignment.query.filter_by(stage_id=stage.id).all()
        assignments[stage.id] = [r.resource.name for r in ass]
    resources = Resource.query.all()
    now = datetime.now().date()
    overdue_in_project = []
    for stage in stages:
        if stage.percent_complete < 100:
            planned_end = dates.get(stage.id, (None, None))[1]
            if planned_end and planned_end < now:
                overdue_in_project.append({
                    'stage': stage,
                    'planned_end': planned_end,
                    'days_overdue': (now - planned_end).days
                })
    predicted_end, deviation_days = predict_completion_date(project, db) if stages else (None, None)
    return render_template('project.html', project=project, stages=stages, dates=dates,
                           assignments=assignments, resources=resources, critical_stages=critical_stages,
                           overdue_in_project=overdue_in_project, predicted_end=predicted_end,
                           deviation_days=deviation_days)

# ========== ПЛАНИРОВАНИЕ ==========
@app.route('/project/<int:project_id>/planning')
@login_required
def project_planning(project_id):
    if not check_project_access(project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    stages = Stage.query.filter_by(project_id=project_id).order_by(Stage.id).all()
    dates = {}
    for s in stages:
        if s.custom_start_date and s.custom_end_date:
            dates[s.id] = (s.custom_start_date, s.custom_end_date)
        else:
            dates.update(calculate_dates(project, db))
    critical_stages = find_critical_path(project, db)
    resources = Resource.query.all()
    assignments = {}
    for stage in stages:
        ass = Assignment.query.filter_by(stage_id=stage.id).all()
        assignments[stage.id] = [r.resource.name for r in ass]
    return render_template('planning.html', project=project, stages=stages, dates=dates,
                           critical_stages=critical_stages, resources=resources, assignments=assignments)

# ========== AJAX ДЛЯ ЭТАПОВ ==========
@app.route('/update_completion_ajax/<int:project_id>', methods=['POST'])
@login_required
def update_completion_ajax(project_id):
    if not check_project_access(project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    data = request.get_json()
    stage_id = data.get('stage_id')
    percent = int(data.get('percent'))
    stage = Stage.query.get_or_404(stage_id)
    if stage.project_id != project_id:
        return jsonify({'success': False, 'error': 'Stage not in project'})
    old_percent = stage.percent_complete
    stage.percent_complete = percent
    if percent >= 100:
        stage.actual_end_date = datetime.now().date()
    else:
        stage.actual_end_date = None
    db.session.commit()
    project = Project.query.get(project_id)
    if percent >= 100 and old_percent < 100:
        create_notification(
            project.user_id,
            f"Этап завершён: {stage.name}",
            f"Этап '{stage.name}' в проекте '{project.name}' выполнен на 100%.",
            url_for('project_view', project_id=project.id)
        )
        for member in project.team:
            if member.user_id != project.user_id:
                create_notification(
                    member.user_id,
                    f"Этап завершён: {stage.name}",
                    f"В проекте '{project.name}' завершён этап '{stage.name}'.",
                    url_for('project_view', project_id=project.id)
                )
    return jsonify({'success': True})

@app.route('/update_stage_date/<int:project_id>', methods=['POST'])
@login_required
def update_stage_date(project_id):
    if not check_project_access(project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    data = request.get_json()
    stage_id = data.get('stage_id')
    field = data.get('field')
    value = data.get('value')
    stage = Stage.query.get_or_404(stage_id)
    if stage.project_id != project_id:
        return jsonify({'success': False, 'error': 'Stage not in project'})
    if field in ('custom_start_date', 'custom_end_date'):
        try:
            new_date = datetime.strptime(value, '%Y-%m-%d').date()
            setattr(stage, field, new_date)
            db.session.commit()
            project = Project.query.get(project_id)
            create_notification(
                project.user_id,
                f"Изменена дата этапа: {stage.name}",
                f"Дата {'начала' if field == 'custom_start_date' else 'окончания'} этапа '{stage.name}' изменена на {new_date.strftime('%d.%m.%Y')}.",
                url_for('project_view', project_id=project_id)
            )
            return jsonify({'success': True})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)})
    return jsonify({'success': False, 'error': 'Invalid field'})

@app.route('/project/<int:project_id>/recalculate_schedule', methods=['POST'])
@login_required
def recalculate_schedule(project_id):
    if not check_project_access(project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    project = Project.query.get_or_404(project_id)
    data = request.get_json()
    changed_stages = data.get('stages', [])
    full_recalc = data.get('full_recalc', False)

    if full_recalc:
        for stage in project.stages:
            if stage.percent_complete < 100:
                stage.custom_start_date = None
                stage.custom_end_date = None
    else:
        for item in changed_stages:
            stage = Stage.query.get(item['id'])
            if stage and stage.project_id == project_id:
                if item.get('start'):
                    try:
                        stage.custom_start_date = datetime.strptime(item['start'], '%Y-%m-%d').date()
                    except:
                        pass
                if item.get('end'):
                    try:
                        stage.custom_end_date = datetime.strptime(item['end'], '%Y-%m-%d').date()
                    except:
                        pass
    db.session.commit()

    new_dates = calculate_dates(project, db)
    for stage in project.stages:
        if stage.id in new_dates:
            stage.custom_start_date = new_dates[stage.id][0]
            stage.custom_end_date = new_dates[stage.id][1]
    db.session.commit()

    dates_json = {str(k): [v[0].isoformat(), v[1].isoformat()] for k, v in new_dates.items()}
    return jsonify({'success': True, 'dates': dates_json})

@app.route('/project/<int:project_id>/shift_stage/<int:stage_id>', methods=['POST'])
@login_required
def shift_stage(project_id, stage_id):
    if not check_project_access(project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    project = Project.query.get_or_404(project_id)
    stage = Stage.query.get_or_404(stage_id)
    if stage.project_id != project_id:
        return jsonify({'success': False, 'error': 'Stage not in project'})
    
    data = request.get_json()
    days = int(data.get('days', 0))
    if days == 0:
        return jsonify({'success': False, 'error': 'No shift amount'})
    
    if stage.custom_start_date:
        stage.custom_start_date = stage.custom_start_date + timedelta(days=days)
    else:
        dates = calculate_dates(project, db)
        if stage.id in dates:
            stage.custom_start_date = dates[stage.id][0] + timedelta(days=days)
    
    if stage.custom_end_date:
        stage.custom_end_date = stage.custom_end_date + timedelta(days=days)
    else:
        dates = calculate_dates(project, db)
        if stage.id in dates:
            stage.custom_end_date = dates[stage.id][1] + timedelta(days=days)
    
    db.session.commit()
    
    new_dates = calculate_dates(project, db)
    for s in project.stages:
        if s.id in new_dates:
            s.custom_start_date = new_dates[s.id][0]
            s.custom_end_date = new_dates[s.id][1]
    db.session.commit()
    
    dates_json = {str(k): [v[0].isoformat(), v[1].isoformat()] for k, v in new_dates.items()}
    return jsonify({'success': True, 'dates': dates_json})

@app.route('/project/<int:project_id>/update_stage_budget/<int:stage_id>', methods=['POST'])
@login_required
def update_stage_budget(project_id, stage_id):
    if not check_project_access(project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    stage = Stage.query.get_or_404(stage_id)
    if stage.project_id != project_id:
        return jsonify({'success': False, 'error': 'Stage not in project'})
    
    data = request.get_json()
    try:
        new_material = float(data.get('planned_material_cost', 0))
        new_labor = float(data.get('planned_labor_cost', 0))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Invalid numbers'})
    
    # Если плановые затраты ещё не заданы — фиксируем план и начальный факт
    if stage.planned_material_cost == 0 and stage.planned_labor_cost == 0:
        stage.planned_material_cost = new_material
        stage.planned_labor_cost = new_labor
        stage.actual_material_cost = new_material
        stage.actual_labor_cost = new_labor
    else:
        # Последующие изменения — корректируем только фактические затраты
        current_total = (stage.planned_material_cost + stage.planned_labor_cost +
                         (stage.actual_material_cost or 0) + (stage.actual_labor_cost or 0))
        new_total = new_material + new_labor
        diff = new_total - current_total
        
        total_actual = (stage.actual_material_cost or 0) + (stage.actual_labor_cost or 0)
        if total_actual > 0:
            material_ratio = (stage.actual_material_cost or 0) / total_actual
            labor_ratio = (stage.actual_labor_cost or 0) / total_actual
        else:
            material_ratio = labor_ratio = 0.5
        
        stage.actual_material_cost = max(0, (stage.actual_material_cost or 0) + diff * material_ratio)
        stage.actual_labor_cost = max(0, (stage.actual_labor_cost or 0) + diff * labor_ratio)
    
    db.session.commit()
    
    project = Project.query.get(project_id)
    create_notification(
        project.user_id,
        f"Изменён бюджет этапа: {stage.name}",
        f"Новый общий бюджет: материалы {stage.planned_material_cost + (stage.actual_material_cost or 0):,.0f} ₽, работа {stage.planned_labor_cost + (stage.actual_labor_cost or 0):,.0f} ₽.",
        url_for('project_view', project_id=project_id)
    )
    
    return jsonify({'success': True})

@app.route('/get_project_dates/<int:project_id>')
@login_required
def get_project_dates(project_id):
    if not check_project_access(project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    project = Project.query.get_or_404(project_id)
    dates = {}
    for s in project.stages:
        if s.custom_start_date and s.custom_end_date:
            dates[s.id] = (s.custom_start_date, s.custom_end_date)
        else:
            dates.update(calculate_dates(project, db))
    dates_json = {str(k): [v[0].isoformat(), v[1].isoformat()] for k, v in dates.items()}
    return jsonify({'success': True, 'dates': dates_json})

@app.route('/api/project/<int:project_id>/stages')
@login_required
def api_project_stages(project_id):
    if not check_project_access(project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    project = Project.query.get_or_404(project_id)
    stages = Stage.query.filter_by(project_id=project_id).order_by(Stage.id).all()
    dates = {}
    for s in stages:
        if s.custom_start_date and s.custom_end_date:
            dates[s.id] = (s.custom_start_date, s.custom_end_date)
        else:
            dates.update(calculate_dates(project, db))

    items = []
    for s in stages:
        start = dates[s.id][0].isoformat() if s.id in dates else project.start_date.isoformat()
        end = dates[s.id][1].isoformat() if s.id in dates else (project.start_date + timedelta(days=s.planned_duration)).isoformat()
        items.append({
            'id': s.id,
            'content': s.name,
            'start': start,
            'end': end,
            'progress': s.percent_complete,
            'className': 'completed' if s.percent_complete >= 100 else ''
        })
    return jsonify({'success': True, 'stages': items})

@app.route('/get_project_progress/<int:project_id>')
@login_required
def get_project_progress(project_id):
    if not check_project_access(project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    stages = Stage.query.filter_by(project_id=project_id).all()
    total = len(stages)
    completed = len([s for s in stages if s.percent_complete >= 100])
    overall_progress = int(completed / total * 100) if total > 0 else 0
    return jsonify({'success': True, 'overall_progress': overall_progress})

@app.route('/get_project_budget/<int:project_id>')
@login_required
def get_project_budget(project_id):
    if not check_project_access(project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    stages = Stage.query.filter_by(project_id=project_id).all()
    total_planned = sum(s.planned_material_cost + s.planned_labor_cost for s in stages)
    total_actual = sum((s.actual_material_cost or 0) + (s.actual_labor_cost or 0) for s in stages)
    return jsonify({'success': True, 'total_planned': total_planned, 'total_actual': total_actual})

@app.route('/project/<int:project_id>/add_stage', methods=['POST'])
@login_required
def add_stage(project_id):
    if not check_project_access(project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    data = request.get_json()
    name = data['name']
    duration = int(data['duration'])
    planned_material_cost = float(data.get('planned_material_cost', 0))
    planned_labor_cost = float(data.get('planned_labor_cost', 0))
    stage = Stage(name=name, planned_duration=duration, project_id=project_id,
                  planned_material_cost=planned_material_cost, planned_labor_cost=planned_labor_cost,
                  actual_material_cost=planned_material_cost, actual_labor_cost=planned_labor_cost)
    db.session.add(stage)
    db.session.flush()
    depends = data.get('depends', [])
    for dep_id in depends:
        dep = Dependency(stage_id=stage.id, depends_on_stage_id=int(dep_id))
        db.session.add(dep)
    resource_ids = data.get('resources', [])
    for rid in resource_ids:
        assign = Assignment(stage_id=stage.id, resource_id=int(rid))
        db.session.add(assign)
    db.session.commit()
    project = Project.query.get(project_id)
    create_notification(
        project.user_id,
        f"Добавлен этап: {stage.name}",
        f"В проекте '{project.name}' добавлен новый этап '{stage.name}'.",
        url_for('project_view', project_id=project.id)
    )
    return jsonify({'success': True, 'stage_id': stage.id})

@app.route('/project/<int:project_id>/edit_stage/<int:stage_id>', methods=['POST'])
@login_required
def edit_stage(project_id, stage_id):
    if not check_project_access(project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    stage = Stage.query.get_or_404(stage_id)
    if stage.project_id != project_id:
        return jsonify({'success': False, 'error': 'Stage not in project'})
    data = request.get_json()
    stage.name = data['name']
    stage.planned_duration = int(data['duration'])
    # Бюджет при редактировании не трогаем (меняется через update_stage_budget)
    if 'percent_complete' in data:
        stage.percent_complete = int(data['percent_complete'])
        if stage.percent_complete >= 100 and not stage.actual_end_date:
            stage.actual_end_date = datetime.now().date()
    Dependency.query.filter_by(stage_id=stage.id).delete()
    depends = data.get('depends', [])
    for dep_id in depends:
        dep = Dependency(stage_id=stage.id, depends_on_stage_id=int(dep_id))
        db.session.add(dep)
    Assignment.query.filter_by(stage_id=stage.id).delete()
    resource_ids = data.get('resources', [])
    for rid in resource_ids:
        assign = Assignment(stage_id=stage.id, resource_id=int(rid))
        db.session.add(assign)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/project/<int:project_id>/delete_stage/<int:stage_id>', methods=['DELETE'])
@login_required
def delete_stage(project_id, stage_id):
    if not check_project_access(project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    stage = Stage.query.get_or_404(stage_id)
    stage_name = stage.name
    if stage.project_id != project_id:
        return jsonify({'success': False, 'error': 'Stage not in project'})
    db.session.delete(stage)
    db.session.commit()
    project = Project.query.get(project_id)
    create_notification(
        project.user_id,
        f"Удалён этап: {stage_name}",
        f"Из проекта '{project.name}' удалён этап '{stage_name}'.",
        url_for('project_view', project_id=project.id)
    )
    return jsonify({'success': True})

# ========== ФОТООТЧЁТЫ ==========
@app.route('/photos/<int:project_id>')
@login_required
def photos_view(project_id):
    if not check_project_access(project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    photos = Photo.query.filter_by(project_id=project_id).order_by(Photo.created_at.desc()).all()
    return render_template('photos.html', project=project, photos=photos)

@app.route('/project/<int:project_id>/upload_photo', methods=['POST'])
@login_required
def upload_photo(project_id):
    if not check_project_access(project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    stage_name = request.form.get('stage_name')
    caption = request.form.get('caption', '')
    file = request.files.get('photo')
    if not file or file.filename == '':
        flash('Файл не выбран', 'error')
        return redirect(url_for('photos_view', project_id=project_id))
    if not stage_name:
        flash('Выберите этап', 'error')
        return redirect(url_for('photos_view', project_id=project_id))
    
    filename = secure_filename(f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
    filepath = os.path.join('uploads', filename)
    full_path = os.path.join('static', filepath)
    file.save(full_path)
    
    photo = Photo(
        project_id=project_id,
        stage_name=stage_name,
        caption=caption,
        filename=filename,
        filepath=filepath.replace('\\', '/')
    )
    db.session.add(photo)
    db.session.commit()
    flash('Фото добавлено', 'success')
    return redirect(url_for('photos_view', project_id=project_id))

# ========== РЕСУРСЫ ==========
@app.route('/resources')
@login_required
def resources_list():
    resources = Resource.query.all()
    return render_template('resources.html', resources=resources, today=date.today())

@app.route('/resources/add', methods=['POST'])
@login_required
def add_resource():
    name = request.form['name']
    if name:
        resource = Resource(name=name)
        db.session.add(resource)
        db.session.commit()
        flash('Ресурс добавлен', 'success')
    return redirect(url_for('resources_list'))

@app.route('/resources/delete/<int:resource_id>')
@login_required
def delete_resource(resource_id):
    resource = Resource.query.get_or_404(resource_id)
    db.session.delete(resource)
    db.session.commit()
    flash('Ресурс удалён', 'success')
    return redirect(url_for('resources_list'))

# ========== ИМПОРТ / ЭКСПОРТ ==========
@app.route('/export_project/<int:project_id>')
@login_required
def export_project(project_id):
    if not check_project_access(project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    stages = Stage.query.filter_by(project_id=project_id).all()
    dates = {}
    for s in stages:
        if s.custom_start_date and s.custom_end_date:
            dates[s.id] = (s.custom_start_date, s.custom_end_date)
        else:
            dates.update(calculate_dates(project, db))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "План работ"
    headers = ['Название', 'Длительность', 'Зависимости', 'Ресурсы', 'План начало', 'План конец', 'Выполнено %', 'Факт завершение', 'План материалы', 'План работы', 'Факт материалы', 'Факт работы']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    for row, stage in enumerate(stages, 2):
        deps = [dep.depends_on_stage.name for dep in stage.dependencies.all()]
        resources = [ass.resource.name for ass in stage.assignments]
        ws.cell(row=row, column=1, value=stage.name)
        ws.cell(row=row, column=2, value=stage.planned_duration)
        ws.cell(row=row, column=3, value=', '.join(deps))
        ws.cell(row=row, column=4, value=', '.join(resources))
        if dates.get(stage.id):
            ws.cell(row=row, column=5, value=dates[stage.id][0].strftime('%Y-%m-%d'))
            ws.cell(row=row, column=6, value=dates[stage.id][1].strftime('%Y-%m-%d'))
        ws.cell(row=row, column=7, value=stage.percent_complete)
        if stage.actual_end_date:
            ws.cell(row=row, column=8, value=stage.actual_end_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=9, value=stage.planned_material_cost)
        ws.cell(row=row, column=10, value=stage.planned_labor_cost)
        ws.cell(row=row, column=11, value=stage.actual_material_cost or 0)
        ws.cell(row=row, column=12, value=stage.actual_labor_cost or 0)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'project_{project.name}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export_csv/<int:project_id>')
@login_required
def export_csv(project_id):
    if not check_project_access(project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    stages = Stage.query.filter_by(project_id=project_id).all()
    dates = {}
    for s in stages:
        if s.custom_start_date and s.custom_end_date:
            dates[s.id] = (s.custom_start_date, s.custom_end_date)
        else:
            dates.update(calculate_dates(project, db))
    si = StringIO()
    writer = csv.writer(si, delimiter=';')
    writer.writerow(['Название этапа', 'Длительность (дни)', 'Зависимости', 'Ресурсы', 'План начало', 'План конец', 'Выполнено %', 'План материалы (руб)', 'План работы (руб)', 'Факт материалы (руб)', 'Факт работы (руб)'])
    for stage in stages:
        deps = [dep.depends_on_stage.name for dep in stage.dependencies.all()]
        resources = [ass.resource.name for ass in stage.assignments]
        writer.writerow([
            stage.name, stage.planned_duration, ', '.join(deps), ', '.join(resources),
            dates[stage.id][0].strftime('%d.%m.%Y') if dates.get(stage.id) else '',
            dates[stage.id][1].strftime('%d.%m.%Y') if dates.get(stage.id) else '',
            stage.percent_complete, stage.planned_material_cost, stage.planned_labor_cost,
            stage.actual_material_cost or 0, stage.actual_labor_cost or 0
        ])
    output = BytesIO()
    output.write(si.getvalue().encode('utf-8-sig'))
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'project_{project.name}.csv', mimetype='text/csv')

@app.route('/project/<int:project_id>/import_stages', methods=['GET', 'POST'])
@login_required
def import_stages(project_id):
    if not check_project_access(project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Файл не выбран', 'error')
            return redirect(url_for('import_stages', project_id=project_id))
        file = request.files['file']
        if file.filename == '':
            flash('Файл не выбран', 'error')
            return redirect(url_for('import_stages', project_id=project_id))
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Поддерживаются только файлы Excel', 'error')
            return redirect(url_for('import_stages', project_id=project_id))
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            imported_count = 0
            skipped_count = 0
            errors = []
            existing_stages = {s.name.strip().lower(): s.id for s in Stage.query.filter_by(project_id=project_id).all()}
            def safe_str(val):
                return str(val).strip() if val is not None else ''
            def parse_float(val):
                if val is None: return 0.0
                if isinstance(val, (int, float)): return float(val)
                s = str(val).strip().replace(' ', '').replace('\xa0', '').replace(',', '.')
                if s == '': return 0.0
                try: return float(s)
                except: raise ValueError(f"Не удалось преобразовать в число: '{val}'")
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                try:
                    name = safe_str(row[0]) if len(row) > 0 else ''
                    if not name:
                        errors.append(f"Строка {row_idx}: пустое название")
                        skipped_count += 1
                        continue
                    duration = int(row[1]) if len(row) > 1 and row[1] is not None else 7
                    depends_on_str = safe_str(row[2]) if len(row) > 2 else ''
                    resources_str = safe_str(row[3]) if len(row) > 3 else ''
                    percent_complete = int(row[4]) if len(row) > 4 and row[4] is not None else 0
                    planned_material = parse_float(row[5]) if len(row) > 5 else 0.0
                    planned_labor = parse_float(row[6]) if len(row) > 6 else 0.0
                    if name.lower() in existing_stages:
                        errors.append(f"Строка {row_idx}: этап '{name}' уже существует")
                        skipped_count += 1
                        continue
                    custom_start = None
                    custom_end = None
                    if len(row) > 7 and row[7] is not None:
                        val = row[7]
                        if isinstance(val, datetime): custom_start = val.date()
                        else:
                            try: custom_start = datetime.strptime(str(val), '%d.%m.%Y').date()
                            except: errors.append(f"Строка {row_idx}: неверный формат даты начала")
                    if len(row) > 8 and row[8] is not None:
                        val = row[8]
                        if isinstance(val, datetime): custom_end = val.date()
                        else:
                            try: custom_end = datetime.strptime(str(val), '%d.%m.%Y').date()
                            except: errors.append(f"Строка {row_idx}: неверный формат даты окончания")
                    stage = Stage(
                        name=name, planned_duration=duration, project_id=project_id,
                        percent_complete=percent_complete, planned_material_cost=planned_material,
                        planned_labor_cost=planned_labor, custom_start_date=custom_start,
                        custom_end_date=custom_end,
                        actual_material_cost=planned_material, actual_labor_cost=planned_labor
                    )
                    db.session.add(stage)
                    db.session.flush()
                    existing_stages[name.lower()] = stage.id
                    if depends_on_str:
                        for dep_name in depends_on_str.split(','):
                            dep_name = dep_name.strip()
                            if dep_name.lower() in existing_stages:
                                dep = Dependency(stage_id=stage.id, depends_on_stage_id=existing_stages[dep_name.lower()])
                                db.session.add(dep)
                            else:
                                errors.append(f"Строка {row_idx}: зависимость '{dep_name}' не найдена")
                    if resources_str:
                        for r_name in resources_str.split(','):
                            r_name = r_name.strip()
                            if not r_name: continue
                            resource = Resource.query.filter_by(name=r_name).first()
                            if not resource:
                                resource = Resource(name=r_name)
                                db.session.add(resource)
                                db.session.flush()
                            assign = Assignment(stage_id=stage.id, resource_id=resource.id)
                            db.session.add(assign)
                    imported_count += 1
                except Exception as e:
                    errors.append(f"Строка {row_idx}: {str(e)}")
                    skipped_count += 1
            db.session.commit()
            flash(f'✅ Импорт завершён: добавлено {imported_count} этапов, пропущено {skipped_count}', 'success')
            if errors:
                for err in errors[:15]:
                    flash(err, 'error')
        except Exception as e:
            flash(f'Ошибка обработки файла: {str(e)}', 'error')
        return redirect(url_for('project_view', project_id=project_id))
    return render_template('import_stages.html', project=project)

# ========== ГЕНЕРАЦИЯ СМЕТЫ ==========
@app.route('/api/generate_estimate_by_budget', methods=['POST'])
@login_required
def generate_estimate_by_budget():
    data = request.get_json()
    budget = data.get('budget')
    quality = data.get('quality', 'standard')
    house_area = data.get('house_area')
    land_area = data.get('land_area')
    if not budget or budget < 50000:
        return jsonify({'success': False, 'error': 'Бюджет должен быть не менее 50 000 ₽'})
    try:
        house_area = float(house_area) if house_area else None
        land_area = float(land_area) if land_area else None
    except: pass
    stages = generate_stages_by_budget(budget, quality, house_area, land_area)
    return jsonify({'success': True, 'stages': stages})

@app.route('/api/add_estimate_stages/<int:project_id>', methods=['POST'])
@login_required
def add_estimate_stages(project_id):
    if not check_project_access(project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    project = Project.query.get_or_404(project_id)
    data = request.get_json()
    stages_data = data.get('stages', [])
    added = 0
    existing_stages = {s.name: s.id for s in Stage.query.filter_by(project_id=project_id).all()}
    for stage_data in stages_data:
        name = stage_data.get('name')
        if not name or name in existing_stages: continue
        duration = stage_data.get('duration', 7)
        budget = stage_data.get('budget', 0)
        custom_start_date = None
        custom_end_date = None
        start_date_str = stage_data.get('startDate')
        end_date_str = stage_data.get('endDate')
        if start_date_str:
            try: custom_start_date = datetime.fromisoformat(start_date_str.replace('Z', '+00:00')).date()
            except: pass
        if end_date_str:
            try: custom_end_date = datetime.fromisoformat(end_date_str.replace('Z', '+00:00')).date()
            except: pass
        mat_cost = int(budget * 0.6)
        lab_cost = int(budget * 0.4)
        stage = Stage(
            name=name, planned_duration=duration, project_id=project_id,
            planned_material_cost=mat_cost, planned_labor_cost=lab_cost,
            percent_complete=0, custom_start_date=custom_start_date, custom_end_date=custom_end_date,
            actual_material_cost=mat_cost, actual_labor_cost=lab_cost
        )
        db.session.add(stage)
        db.session.flush()
        existing_stages[name] = stage.id
        added += 1
        resources = stage_data.get('resources', '')
        if resources and resources != '—':
            for r_name in resources.split(', '):
                resource = Resource.query.filter_by(name=r_name).first()
                if not resource:
                    resource = Resource(name=r_name)
                    db.session.add(resource)
                    db.session.flush()
                assign = Assignment(stage_id=stage.id, resource_id=resource.id)
                db.session.add(assign)
        depends_on = stage_data.get('depends_on')
        if depends_on and depends_on != '—' and depends_on in existing_stages:
            dep = Dependency(stage_id=stage.id, depends_on_stage_id=existing_stages[depends_on])
            db.session.add(dep)
    db.session.commit()
    create_notification(project.user_id, "Смета добавлена", f"В проект '{project.name}' добавлено {added} этапов.", url_for('project_view', project_id=project.id))
    return jsonify({'success': True, 'count': added})

# ========== ПРОГНОЗ ==========
@app.route('/get_project_forecast/<int:project_id>')
@login_required
def get_project_forecast(project_id):
    if not check_project_access(project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    project = Project.query.get_or_404(project_id)
    predicted_end, deviation = predict_completion_date(project, db)
    return jsonify({'success': True, 'predicted_end': predicted_end.isoformat() if predicted_end else None, 'deviation': deviation})

# ========== ШАБЛОНЫ ПРОЕКТОВ ==========
@app.route('/templates')
@login_required
def templates_list():
    templates = ProjectTemplate.query.filter_by(user_id=current_user.id).all()
    return render_template('templates.html', templates=templates)

@app.route('/templates/save/<int:project_id>', methods=['POST'])
@login_required
def save_as_template(project_id):
    if not check_project_access(project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    name = request.form.get('name')
    description = request.form.get('description', '')
    if not name:
        flash('Введите название шаблона', 'error')
        return redirect(url_for('project_view', project_id=project_id))
    stages = Stage.query.filter_by(project_id=project_id).all()
    stages_list = []
    for stage in stages:
        deps = [dep.depends_on_stage_id for dep in stage.dependencies.all()]
        resources = [ass.resource.name for ass in stage.assignments]
        stages_list.append({
            'name': stage.name, 'duration': stage.planned_duration,
            'dependencies': deps, 'resources': resources,
            'planned_material_cost': stage.planned_material_cost,
            'planned_labor_cost': stage.planned_labor_cost
        })
    template = ProjectTemplate(user_id=current_user.id, name=name, description=description, stages_data=json.dumps(stages_list))
    db.session.add(template)
    db.session.commit()
    flash(f'Шаблон "{name}" сохранён', 'success')
    return redirect(url_for('templates_list'))

@app.route('/templates/delete/<int:template_id>')
@login_required
def delete_template(template_id):
    template = ProjectTemplate.query.get_or_404(template_id)
    if template.user_id != current_user.id and not current_user.is_admin():
        flash('Доступ запрещён', 'error')
        return redirect(url_for('templates_list'))
    db.session.delete(template)
    db.session.commit()
    flash('Шаблон удалён', 'success')
    return redirect(url_for('templates_list'))

@app.route('/project/<int:project_id>/apply_template/<int:template_id>')
@login_required
def apply_template(project_id, template_id):
    if not check_project_access(project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    template = ProjectTemplate.query.get_or_404(template_id)
    if template.user_id != current_user.id and not current_user.is_admin():
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    stages_data = json.loads(template.stages_data)
    name_to_id = {}
    for stage_info in stages_data:
        stage = Stage(
            name=stage_info['name'], planned_duration=stage_info['duration'], project_id=project_id,
            planned_material_cost=stage_info.get('planned_material_cost', 0),
            planned_labor_cost=stage_info.get('planned_labor_cost', 0),
            actual_material_cost=stage_info.get('planned_material_cost', 0),
            actual_labor_cost=stage_info.get('planned_labor_cost', 0)
        )
        db.session.add(stage)
        db.session.flush()
        name_to_id[stage_info['name']] = stage.id
        for r_name in stage_info.get('resources', []):
            resource = Resource.query.filter_by(name=r_name).first()
            if not resource:
                resource = Resource(name=r_name)
                db.session.add(resource)
                db.session.flush()
            assign = Assignment(stage_id=stage.id, resource_id=resource.id)
            db.session.add(assign)
    for stage_info in stages_data:
        stage_id = name_to_id[stage_info['name']]
        for dep_name in stage_info.get('dependencies', []):
            if dep_name in name_to_id:
                dep = Dependency(stage_id=stage_id, depends_on_stage_id=name_to_id[dep_name])
                db.session.add(dep)
    db.session.commit()
    flash(f'Этапы из шаблона "{template.name}" добавлены', 'success')
    return redirect(url_for('project_view', project_id=project_id))

@app.route('/create_from_template/<int:template_id>', methods=['POST'])
@login_required
def create_from_template(template_id):
    template = ProjectTemplate.query.get_or_404(template_id)
    if template.user_id != current_user.id and not current_user.is_admin():
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    name = request.form['name']
    start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date()
    address = request.form.get('address', '')
    try:
        lat = float(request.form.get('latitude', 55.751244))
        lon = float(request.form.get('longitude', 37.618423))
    except:
        lat, lon = 55.751244, 37.618423
    project = Project(name=name, start_date=start_date, address=address, latitude=lat, longitude=lon, user_id=current_user.id)
    db.session.add(project)
    db.session.flush()
    stages_data = json.loads(template.stages_data)
    name_to_id = {}
    for stage_info in stages_data:
        stage = Stage(
            name=stage_info['name'], planned_duration=stage_info['duration'], project_id=project.id,
            planned_material_cost=stage_info.get('planned_material_cost', 0),
            planned_labor_cost=stage_info.get('planned_labor_cost', 0),
            actual_material_cost=stage_info.get('planned_material_cost', 0),
            actual_labor_cost=stage_info.get('planned_labor_cost', 0)
        )
        db.session.add(stage)
        db.session.flush()
        name_to_id[stage_info['name']] = stage.id
        for r_name in stage_info.get('resources', []):
            resource = Resource.query.filter_by(name=r_name).first()
            if not resource:
                resource = Resource(name=r_name)
                db.session.add(resource)
                db.session.flush()
            assign = Assignment(stage_id=stage.id, resource_id=resource.id)
            db.session.add(assign)
    for stage_info in stages_data:
        stage_id = name_to_id[stage_info['name']]
        for dep_name in stage_info.get('dependencies', []):
            if dep_name in name_to_id:
                dep = Dependency(stage_id=stage_id, depends_on_stage_id=name_to_id[dep_name])
                db.session.add(dep)
    db.session.commit()
    flash(f'Проект "{name}" создан из шаблона "{template.name}"', 'success')
    return redirect(url_for('project_view', project_id=project.id))

# ========== ДАШБОРДЫ ==========
@app.route('/dashboard')
@login_required
def dashboard():
    owned = Project.query.filter_by(user_id=current_user.id).all()
    team_projects = db.session.query(Project).join(ProjectTeam).filter(ProjectTeam.user_id == current_user.id).all()
    projects = list(set(owned + team_projects))
    stats = []
    overdue_stages = get_overdue_stages()
    overdue_stages = [os for os in overdue_stages if check_project_access(os['project'].id)]
    selected_project_id = request.args.get('project_id', type=int)
    selected_project = None
    weather = None
    if selected_project_id and check_project_access(selected_project_id):
        selected_project = Project.query.get(selected_project_id)
        if selected_project:
            weather = get_weather_by_coords(selected_project.latitude, selected_project.longitude)
    for project in projects:
        stages = Stage.query.filter_by(project_id=project.id).all()
        total_stages = len(stages)
        completed_stages = len([s for s in stages if s.percent_complete >= 100])
        delayed = 0
        for stage in stages:
            if stage.percent_complete < 100:
                if stage.custom_end_date:
                    planned_end = stage.custom_end_date
                else:
                    dates = calculate_dates(project, db)
                    planned_end = dates.get(stage.id, (None, None))[1]
                if planned_end and planned_end < datetime.now().date():
                    delayed += 1
        total_planned = sum(s.planned_material_cost + s.planned_labor_cost for s in stages)
        total_actual = sum((s.actual_material_cost or 0) + (s.actual_labor_cost or 0) for s in stages)
        stats.append({
            'project': project, 'total_stages': total_stages, 'completed_stages': completed_stages,
            'delayed_stages': delayed, 'progress': int(completed_stages / total_stages * 100) if total_stages > 0 else 0,
            'budget': {'planned': total_planned, 'actual': total_actual, 'diff': total_actual - total_planned}
        })
    return render_template('dashboard.html', stats=stats, overdue_stages=overdue_stages, weather=weather, selected_project=selected_project, projects=projects)

@app.route('/executive_dashboard')
@login_required
def executive_dashboard():
    owned = Project.query.filter_by(user_id=current_user.id).all()
    team_projects = db.session.query(Project).join(ProjectTeam).filter(ProjectTeam.user_id == current_user.id).all()
    projects = list(set(owned + team_projects))
    total_budget = 0
    total_spent = 0
    total_stages = 0
    completed_stages = 0
    overdue_stages_count = 0
    project_stats = []
    all_stages_info = []
    for project in projects:
        stages = Stage.query.filter_by(project_id=project.id).all()
        dates = {}
        for s in stages:
            if s.custom_start_date and s.custom_end_date:
                dates[s.id] = (s.custom_start_date, s.custom_end_date)
            else:
                dates.update(calculate_dates(project, db))
        project_budget = sum(s.planned_material_cost + s.planned_labor_cost for s in stages)
        project_spent = sum((s.actual_material_cost or 0) + (s.actual_labor_cost or 0) for s in stages)
        project_total_stages = len(stages)
        project_completed = len([s for s in stages if s.percent_complete >= 100])
        overdue = 0
        for stage in stages:
            if stage.percent_complete < 100:
                planned_end = dates.get(stage.id, (None, None))[1]
                if planned_end and planned_end < datetime.now().date():
                    overdue += 1
                start_date = dates.get(stage.id, (None, None))[0]
                end_date = planned_end
                days_deviation = None
                if end_date and start_date:
                    expected_end = start_date + timedelta(days=stage.planned_duration - 1)
                    days_deviation = (end_date - expected_end).days
                all_stages_info.append({
                    'project': project, 'stage': stage, 'start_date': start_date, 'end_date': end_date,
                    'is_overdue': planned_end and planned_end < datetime.now().date(),
                    'is_upcoming': planned_end and datetime.now().date() <= planned_end <= datetime.now().date() + timedelta(days=7),
                    'days_deviation': days_deviation
                })
        total_budget += project_budget
        total_spent += project_spent
        total_stages += project_total_stages
        completed_stages += project_completed
        overdue_stages_count += overdue
        completion_date = None
        if stages:
            ends = [s.custom_end_date for s in stages if s.custom_end_date]
            if ends: completion_date = max(ends)
            else:
                ends_calc = [dates[s.id][1] for s in stages if s.id in dates]
                if ends_calc: completion_date = max(ends_calc)
        project_stats.append({
            'project': project, 'budget': project_budget, 'spent': project_spent,
            'total_stages': project_total_stages, 'completed_stages': project_completed,
            'overdue': overdue, 'completion_date': completion_date,
            'progress': int(project_completed / project_total_stages * 100) if project_total_stages > 0 else 0,
            'budget_usage': int(project_spent / project_budget * 100) if project_budget > 0 else 0
        })
    overall_progress = int(completed_stages / total_stages * 100) if total_stages > 0 else 0
    budget_usage = int(total_spent / total_budget * 100) if total_budget > 0 else 0
    overdue_stages = get_overdue_stages()
    overdue_stages = [os for os in overdue_stages if check_project_access(os['project'].id)]
    return render_template('executive_dashboard.html',
                           projects=projects, project_stats=project_stats, total_projects=len(projects),
                           total_budget=total_budget, total_spent=total_spent, total_stages=total_stages,
                           completed_stages=completed_stages, overdue_stages=overdue_stages,
                           overdue_stages_count=overdue_stages_count, overall_progress=overall_progress,
                           budget_usage=budget_usage, all_stages_info=all_stages_info)

# ========== ДОГОВОРЫ ==========
@app.route('/contracts/<int:project_id>')
@login_required
def contracts(project_id):
    if not check_project_access(project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    contracts = Contract.query.filter_by(project_id=project_id).order_by(Contract.contract_date.desc()).all()
    return render_template('contracts.html', project=project, contracts=contracts)

@app.route('/add_contract/<int:project_id>', methods=['GET', 'POST'])
@login_required
def add_contract(project_id):
    if not check_project_access(project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    if request.method == 'POST':
        contract = Contract(
            project_id=project_id, contract_number=request.form['contract_number'],
            contract_date=datetime.strptime(request.form['contract_date'], '%Y-%m-%d').date(),
            contractor_name=request.form['contractor_name'],
            contractor_inn=request.form.get('contractor_inn', ''),
            contractor_kpp=request.form.get('contractor_kpp', ''),
            contractor_ogrn=request.form.get('contractor_ogrn', ''),
            contractor_address=request.form.get('contractor_address', ''),
            contractor_fact_address=request.form.get('contractor_fact_address', ''),
            contractor_phone=request.form.get('contractor_phone', ''),
            contractor_email=request.form.get('contractor_email', ''),
            contractor_director=request.form.get('contractor_director', ''),
            contractor_basis=request.form.get('contractor_basis', ''),
            contractor_bank_name=request.form.get('contractor_bank_name', ''),
            contractor_bik=request.form.get('contractor_bik', ''),
            contractor_account=request.form.get('contractor_account', ''),
            contractor_correspondent_account=request.form.get('contractor_correspondent_account', ''),
            start_date=datetime.strptime(request.form['start_date'], '%Y-%m-%d').date(),
            end_date=datetime.strptime(request.form['end_date'], '%Y-%m-%d').date() if request.form.get('end_date') else None,
            total_amount=float(request.form['total_amount']),
            object_name=request.form.get('object_name', project.name),
            object_address=request.form.get('object_address', project.address),
            subject=request.form.get('subject', ''), documentation=request.form.get('documentation', ''),
            doc_deadline=request.form.get('doc_deadline', ''), doc_method=request.form.get('doc_method', ''),
            special_conditions=request.form.get('special_conditions', ''),
            liability=request.form.get('liability', ''), warranty=request.form.get('warranty', ''),
            status='draft'
        )
        db.session.add(contract)
        db.session.flush()
        work_names = request.form.getlist('work_name[]')
        work_units = request.form.getlist('work_unit[]')
        work_qtys = request.form.getlist('work_qty[]')
        work_prices = request.form.getlist('work_price[]')
        work_totals = request.form.getlist('work_total[]')
        for i in range(len(work_names)):
            if work_names[i] and work_names[i].strip():
                work_item = ContractWorkItem(
                    contract_id=contract.id, work_name=work_names[i],
                    unit=work_units[i] if i < len(work_units) else 'шт',
                    quantity=float(work_qtys[i]) if i < len(work_qtys) and work_qtys[i] else 0,
                    unit_price=float(work_prices[i]) if i < len(work_prices) and work_prices[i] else 0,
                    total_price=float(work_totals[i]) if i < len(work_totals) and work_totals[i] else 0
                )
                db.session.add(work_item)
        db.session.commit()
        flash('Договор добавлен', 'success')
        return redirect(url_for('contracts', project_id=project_id))
    return render_template('add_contract.html', project=project)

@app.route('/sign_contract/<int:contract_id>')
@login_required
def sign_contract(contract_id):
    contract = Contract.query.get_or_404(contract_id)
    if not check_project_access(contract.project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    contract.status = 'signed'
    db.session.commit()
    flash('Договор подписан', 'success')
    return redirect(url_for('contracts', project_id=contract.project_id))

@app.route('/delete_contract/<int:contract_id>')
@login_required
def delete_contract(contract_id):
    contract = Contract.query.get_or_404(contract_id)
    if not check_project_access(contract.project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project_id = contract.project_id
    db.session.delete(contract)
    db.session.commit()
    flash('Договор удалён', 'success')
    return redirect(url_for('contracts', project_id=project_id))

# ========== ПЛАН ЗАКУПОК ==========
@app.route('/purchase_plan/<int:project_id>')
@login_required
def purchase_plan(project_id):
    if not check_project_access(project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    purchase_orders = PurchaseOrder.query.filter_by(project_id=project_id).order_by(PurchaseOrder.order_date.desc()).all()
    return render_template('purchase_plan.html', project=project, purchase_orders=purchase_orders)

@app.route('/generate_purchase_plan/<int:project_id>')
@login_required
def generate_purchase_plan(project_id):
    if not check_project_access(project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    stages = Stage.query.filter_by(project_id=project_id).all()
    dates = {}
    for s in stages:
        if s.custom_start_date and s.custom_end_date:
            dates[s.id] = (s.custom_start_date, s.custom_end_date)
        else:
            dates.update(calculate_dates(project, db))
    materials_by_category = {}
    for stage in stages:
        if stage.planned_material_cost > 0 and dates.get(stage.id):
            materials = get_materials_for_stage(stage.name, stage.planned_material_cost)
            for material in materials:
                category = material['supplier'] or 'Прочие'
                if category not in materials_by_category:
                    materials_by_category[category] = []
                materials_by_category[category].append({
                    'stage': stage, 'material_name': material['name'],
                    'quantity': material['quantity'], 'unit': material['unit'],
                    'unit_price': material['unit_price'], 'total_price': material['total_price'],
                    'delivery_date': dates[stage.id][0] - timedelta(days=7)
                })
    idx = 0
    for supplier, items in materials_by_category.items():
        idx += 1
        order_number = f"PO-{project_id}-{datetime.now().strftime('%Y%m%d')}-{idx}"
        order = PurchaseOrder(
            project_id=project_id, order_number=order_number,
            supplier=supplier if supplier != 'Прочие' else None,
            order_date=datetime.now().date(),
            total_amount=sum(item['total_price'] for item in items),
            status='pending'
        )
        db.session.add(order)
        db.session.flush()
        for item in items:
            purchase_item = PurchaseItem(
                order_id=order.id, stage_id=item['stage'].id,
                material_name=item['material_name'], quantity=item['quantity'],
                unit=item['unit'], unit_price=item['unit_price'],
                total_price=item['total_price'], status='pending'
            )
            db.session.add(purchase_item)
    db.session.commit()
    flash('План закупок сформирован!', 'success')
    return redirect(url_for('purchase_plan', project_id=project_id))

@app.route('/project/<int:project_id>/add_order', methods=['POST'])
@login_required
def add_purchase_order(project_id):
    if not check_project_access(project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project = Project.query.get_or_404(project_id)
    order_number = request.form['order_number']
    supplier = request.form.get('supplier', '')
    order_date = datetime.strptime(request.form['order_date'], '%Y-%m-%d').date()
    order = PurchaseOrder(project_id=project_id, order_number=order_number, supplier=supplier, order_date=order_date, status='pending')
    db.session.add(order)
    db.session.commit()
    flash('Заказ создан', 'success')
    return redirect(url_for('purchase_plan', project_id=project_id))

@app.route('/order/<int:order_id>/add_item', methods=['POST'])
@login_required
def add_order_item(order_id):
    order = PurchaseOrder.query.get_or_404(order_id)
    if not check_project_access(order.project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    data = request.get_json()
    stage_id = data.get('stage_id')
    material_name = data.get('material_name')
    quantity = float(data.get('quantity', 0))
    unit = data.get('unit', 'шт')
    unit_price = float(data.get('unit_price', 0))
    total_price = quantity * unit_price
    item = PurchaseItem(order_id=order_id, stage_id=stage_id, material_name=material_name, quantity=quantity, unit=unit, unit_price=unit_price, total_price=total_price, status='pending')
    db.session.add(item)
    order.total_amount = sum(i.total_price for i in order.items) + total_price
    db.session.commit()
    return jsonify({'success': True, 'item_id': item.id})

@app.route('/order/item/<int:item_id>/update', methods=['POST'])
@login_required
def update_order_item(item_id):
    item = PurchaseItem.query.get_or_404(item_id)
    order = item.order
    if not check_project_access(order.project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    data = request.get_json()
    item.material_name = data.get('material_name', item.material_name)
    item.quantity = float(data.get('quantity', item.quantity))
    item.unit = data.get('unit', item.unit)
    item.unit_price = float(data.get('unit_price', item.unit_price))
    item.total_price = item.quantity * item.unit_price
    db.session.commit()
    order.total_amount = sum(i.total_price for i in order.items)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/order/item/<int:item_id>/delete', methods=['DELETE'])
@login_required
def delete_order_item(item_id):
    item = PurchaseItem.query.get_or_404(item_id)
    order = item.order
    if not check_project_access(order.project_id):
        return jsonify({'success': False, 'error': 'Access denied'})
    db.session.delete(item)
    order.total_amount = sum(i.total_price for i in order.items if i.id != item_id)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/update_order_status/<int:order_id>/<status>')
@login_required
def update_order_status(order_id, status):
    order = PurchaseOrder.query.get_or_404(order_id)
    if not check_project_access(order.project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    order.status = status
    for item in order.items:
        if status == 'ordered': item.status = 'ordered'
        elif status == 'delivered': item.status = 'received'
    db.session.commit()
    flash(f'Статус заказа №{order.order_number} обновлён', 'success')
    return redirect(url_for('purchase_plan', project_id=order.project_id))

@app.route('/delete_order/<int:order_id>')
@login_required
def delete_order(order_id):
    order = PurchaseOrder.query.get_or_404(order_id)
    if not check_project_access(order.project_id):
        flash('Доступ запрещён', 'error')
        return redirect(url_for('index'))
    project_id = order.project_id
    db.session.delete(order)
    db.session.commit()
    flash('Заказ удалён', 'success')
    return redirect(url_for('purchase_plan', project_id=project_id))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)